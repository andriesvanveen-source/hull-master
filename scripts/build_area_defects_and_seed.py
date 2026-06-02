from __future__ import annotations

import csv
import json
import re
import zipfile
from collections import defaultdict
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
COMMON_CSV = ROOT / "public" / "common-defects.csv"
SOURCE_XLSX = ROOT / "C1_CE_Reports (new) (2).xlsx"
AREA_XLSX = ROOT / "area-specific-common-defects.xlsx"
AREA_JSON = ROOT / "public" / "area-common-defects.json"
SEED_SQL = ROOT / "supabase-seed-previous-boats.sql"

APP_AREAS = [
    "Saloon",
    "Aft Cockpit",
    "Fwd Cockpit Lounge & Deck",
    "Stbd Engine",
    "Port Engine",
    "Stbd Fwd Cabin & Heads",
    "Stbd Mid Cabin & Heads",
    "Stbd Aft Cabin & Heads",
    "Port Fwd Cabin & Heads",
    "Port Mid Cabin & Heads",
    "Port Aft Cabin & Heads",
    "Crew Cabin",
]

CSV_AREA_MAP = {
    "Saloon": "Saloon",
    "Aft Cockpit": "Aft Cockpit",
    "Stbd Engine": "Stbd Engine",
    "Stb Engine Room": "Stbd Engine",
    "Port Engine": "Port Engine",
    "Prt Engine Room": "Port Engine",
    "Stbd Fwd": "Stbd Fwd Cabin & Heads",
    "Stbd Mid": "Stbd Mid Cabin & Heads",
    "Stbd Aft": "Stbd Aft Cabin & Heads",
    "Prt Fwd": "Port Fwd Cabin & Heads",
    "Prt Passage": "Port Mid Cabin & Heads",
    "Prt Aft": "Port Aft Cabin & Heads",
    "FWD COCKPIT, SIDE DECKS AND COACH ROOF": "Fwd Cockpit Lounge & Deck",
}

REPORT_AREA_MAP = {
    "SALOON": "Saloon",
    "AFT COCKPIT": "Aft Cockpit",
    "FWD COCKPIT, SIDE DECKS AND COACH ROOF": "Fwd Cockpit Lounge & Deck",
    "STB ENGINE ROOM": "Stbd Engine",
    "STBD ENGINE ROOM": "Stbd Engine",
    "PORT ENGINE ROOM": "Port Engine",
    "STBD FWD CABIN AND HEADS": "Stbd Fwd Cabin & Heads",
    "STBD FWD CABIN & HEADS": "Stbd Fwd Cabin & Heads",
    "STBD MID CABIN AND HEADS": "Stbd Mid Cabin & Heads",
    "STBD MID CABIN & HEADS": "Stbd Mid Cabin & Heads",
    "STBD AFT CABIN AND HEADS": "Stbd Aft Cabin & Heads",
    "STBD AFT CABIN & HEADS": "Stbd Aft Cabin & Heads",
    "PORT FWD CABIN AND HEADS": "Port Fwd Cabin & Heads",
    "PORT FWD CABIN & HEADS": "Port Fwd Cabin & Heads",
    "PORT PASSAGE": "Port Mid Cabin & Heads",
    "PORT MID CABIN AND HEADS": "Port Mid Cabin & Heads",
    "PORT MID CABIN & HEADS": "Port Mid Cabin & Heads",
    "PORT AFT CABIN AND HEADS": "Port Aft Cabin & Heads",
    "PORT AFT CABIN & HEADS": "Port Aft Cabin & Heads",
    "CREW CABIN": "Crew Cabin",
}


def normalize_text(value: str) -> str:
    return re.sub(r"\s+", " ", value.strip()).lower()


def normalize_area_heading(value: str) -> str:
    cleaned = re.sub(r"\s+", " ", value.strip())
    cleaned = cleaned.replace("& Heads", "& HEADS")
    return cleaned.upper()


def sql_literal(value: str | None) -> str:
    if value is None or value == "":
        return "null"
    return "'" + value.replace("'", "''") + "'"


def load_common_defects_from_csv() -> tuple[dict[str, list[dict[str, str]]], dict[str, str]]:
    by_area: dict[str, list[dict[str, str]]] = {area: [] for area in APP_AREAS}
    discipline_by_text: dict[str, str] = {}
    seen_by_area: dict[str, set[str]] = defaultdict(set)

    with COMMON_CSV.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            text = (row.get("Common Defect List") or "").strip()
            discipline = (row.get("Task Allocation") or "").strip().upper()

            if not text:
                continue

            if discipline:
                discipline_by_text[normalize_text(text)] = discipline

            for csv_area, app_area in CSV_AREA_MAP.items():
                if (row.get(csv_area) or "").strip().lower() == "yes":
                    key = normalize_text(text)

                    if key in seen_by_area[app_area]:
                        continue

                    by_area[app_area].append({
                        "text": text,
                        "discipline": discipline,
                    })
                    seen_by_area[app_area].add(key)

    return by_area, discipline_by_text


def load_common_defects_from_area_workbook() -> tuple[dict[str, list[dict[str, str]]], dict[str, str]]:
    by_area: dict[str, list[dict[str, str]]] = {area: [] for area in APP_AREAS}
    discipline_by_text: dict[str, str] = {}
    seen_by_area: dict[str, set[str]] = defaultdict(set)

    if not AREA_XLSX.exists():
        return load_common_defects_from_csv()

    workbook = __import__("openpyxl").load_workbook(AREA_XLSX, read_only=True, data_only=True)

    for area in APP_AREAS:
        if area[:31] not in workbook.sheetnames:
            continue

        sheet = workbook[area[:31]]

        for row in sheet.iter_rows(min_row=2, values_only=True):
            text = str(row[0] or "").strip()
            discipline = str(row[1] or "").strip().upper()

            if not text:
                continue

            key = normalize_text(text)

            if key in seen_by_area[area]:
                continue

            by_area[area].append({
                "text": text,
                "discipline": discipline,
            })
            seen_by_area[area].add(key)

            if discipline:
                discipline_by_text[key] = discipline

    return by_area, discipline_by_text


def build_area_workbook(by_area: dict[str, list[dict[str, str]]]) -> None:
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    summary.append(["Area", "Defect Count"])

    for area in APP_AREAS:
        summary.append([area, len(by_area[area])])

    header_fill = PatternFill("solid", fgColor="0B2D49")
    header_font = Font(color="FFFFFF", bold=True)
    subheader_fill = PatternFill("solid", fgColor="DFE8E3")

    for cell in summary[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    summary.column_dimensions["A"].width = 34
    summary.column_dimensions["B"].width = 16
    summary.freeze_panes = "A2"

    for area in APP_AREAS:
        sheet = wb.create_sheet(area[:31])
        sheet.append(["Defect Description", "Discipline"])

        for defect in by_area[area]:
            sheet.append([defect["text"], defect["discipline"]])

        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row in sheet.iter_rows(min_row=2):
            row[0].alignment = Alignment(wrap_text=True, vertical="top")
            row[1].alignment = Alignment(horizontal="center")

        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:B{max(sheet.max_row, 1)}"
        sheet.column_dimensions["A"].width = 58
        sheet.column_dimensions["B"].width = 14

        if sheet.max_row == 1:
            sheet["A2"] = ""
            sheet["B2"] = ""
            sheet["A2"].fill = subheader_fill

    wb.save(AREA_XLSX)


def save_area_json(by_area: dict[str, list[dict[str, str]]]) -> None:
    AREA_JSON.write_text(
        json.dumps(by_area, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )


def xlsx_shared_strings(zf: zipfile.ZipFile) -> list[str]:
    strings: list[str] = []
    root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
    ns = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}

    for item in root.findall("a:si", ns):
        strings.append("".join(t.text or "" for t in item.iter("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t")))

    return strings


def column_row(ref: str) -> tuple[int, int]:
    match = re.match(r"([A-Z]+)([0-9]+)", ref)
    if not match:
        raise ValueError(f"Invalid cell reference: {ref}")

    col = 0
    for char in match.group(1):
        col = col * 26 + ord(char) - 64

    return int(match.group(2)), col


def read_xlsx_rows(zf: zipfile.ZipFile, sheet_target: str, strings: list[str]) -> dict[int, dict[int, str]]:
    ns = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    target = sheet_target.lstrip("/")

    if not target.startswith("xl/"):
        target = f"xl/{target}"

    root = ET.fromstring(zf.read(target))
    rows: dict[int, dict[int, str]] = {}

    for cell in root.findall(".//a:c", ns):
        ref = cell.attrib.get("r")
        value_node = cell.find("a:v", ns)

        if not ref or value_node is None:
            continue

        row_number, column_number = column_row(ref)
        value = value_node.text or ""

        if cell.attrib.get("t") == "s":
            value = strings[int(value)]

        value = value.strip() if isinstance(value, str) else str(value)

        if value:
            rows.setdefault(row_number, {})[column_number] = value

    return rows


def workbook_sheet_targets(zf: zipfile.ZipFile) -> dict[str, str]:
    main_ns = {
        "a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        "rel": "http://schemas.openxmlformats.org/package/2006/relationships",
    }
    workbook = ET.fromstring(zf.read("xl/workbook.xml"))
    relationships = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
    relationship_targets = {item.attrib["Id"]: item.attrib["Target"] for item in relationships}
    targets: dict[str, str] = {}

    for sheet in workbook.find("a:sheets", main_ns):
        name = sheet.attrib["name"]
        rel_id = sheet.attrib["{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"]
        targets[name] = relationship_targets[rel_id]

    return targets


def extract_previous_boats(discipline_by_text: dict[str, str]) -> dict[str, list[dict[str, str]]]:
    boats: dict[str, list[dict[str, str]]] = {}

    with zipfile.ZipFile(SOURCE_XLSX) as zf:
        strings = xlsx_shared_strings(zf)
        targets = workbook_sheet_targets(zf)

        for sheet_name, target in targets.items():
            if not re.fullmatch(r"C\d{3,5}", sheet_name):
                continue

            rows = read_xlsx_rows(zf, target, strings)
            defects: list[dict[str, str]] = []
            current_area = ""

            for row_number in sorted(rows):
                values = rows[row_number]
                area_candidate = values.get(4, "")
                normalized_area = normalize_area_heading(area_candidate)

                if normalized_area in REPORT_AREA_MAP:
                    current_area = REPORT_AREA_MAP[normalized_area]
                    continue

                defect_text = values.get(4, "")
                number = values.get(3, "")

                if current_area and defect_text and re.fullmatch(r"\d+", number):
                    defects.append({
                        "text": defect_text,
                        "discipline": discipline_by_text.get(normalize_text(defect_text), ""),
                        "area": current_area,
                    })

            boats[sheet_name] = defects

    return boats


def save_seed_sql(boats: dict[str, list[dict[str, str]]]) -> None:
    lines = [
        "-- Seed previous Hull Master boats from C1_CE_Reports (new) (2).xlsx",
        "-- Run this after supabase-schema.sql.",
        "",
    ]

    for boat_name, defects in boats.items():
        lines.append(f"insert into public.boats (name)")
        lines.append(f"values ({sql_literal(boat_name)})")
        lines.append("on conflict (name) do nothing;")
        lines.append("")

        for defect in defects:
            lines.append("insert into public.defects (boat_id, text, discipline, area)")
            lines.append(
                "select b.id, "
                f"{sql_literal(defect['text'])}, "
                f"{sql_literal(defect['discipline'])}, "
                f"{sql_literal(defect['area'])}"
            )
            lines.append("from public.boats b")
            lines.append(f"where b.name = {sql_literal(boat_name)}")
            lines.append("  and not exists (")
            lines.append("    select 1")
            lines.append("    from public.defects d")
            lines.append("    where d.boat_id = b.id")
            lines.append(f"      and d.area = {sql_literal(defect['area'])}")
            lines.append(f"      and lower(btrim(d.text)) = lower(btrim({sql_literal(defect['text'])}))")
            lines.append("  );")
            lines.append("")

    SEED_SQL.write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    by_area, discipline_by_text = load_common_defects_from_area_workbook()
    build_area_workbook(by_area)
    save_area_json(by_area)
    boats = extract_previous_boats(discipline_by_text)
    save_seed_sql(boats)

    total_area_defects = sum(len(items) for items in by_area.values())
    total_boat_defects = sum(len(items) for items in boats.values())
    print(json.dumps({
        "areaWorkbook": str(AREA_XLSX),
        "areaJson": str(AREA_JSON),
        "seedSql": str(SEED_SQL),
        "areaDefects": total_area_defects,
        "boats": len(boats),
        "boatDefects": total_boat_defects,
    }, indent=2))


if __name__ == "__main__":
    main()
